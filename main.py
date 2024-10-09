import pandas
import pandas as pd
import openpyxl
import math
from openpyxl.styles import Font,Alignment, Side, Border, PatternFill
from openpyxl import Workbook, load_workbook
from datetime import datetime

pd.set_option('display.max_columns', None)
sheet1 = pandas.read_excel("待填工時表單-20231225-2-2-2.xlsx")
sheet2 = pandas.read_excel("待填工時表單-20231225-2-2-2.xlsx", sheet_name="第一周上班人數")
waiting_time_sheet = pd.read_excel("待填工時表單-20231225-2-2-2.xlsx", header=1)
#紀錄損耗率
Loss_rate = 'Loss rate.txt'

#========================計算加工時間=======================================
data1 = pandas.read_excel("模擬2023年12月25製令單.xlsx",
                        sheet_name='製令單')      #讀取 Excel,並選擇工作表
data2 = pandas.read_excel('福佑電機製造部工時總攬資料(新).xlsx',
                        sheet_name='產品途程明細表 (主檔) 20190214',
                        header=1)        #將標頭設為第二列



#消除空白字元
data1['製令編號']=data1['製令編號'].astype(str).replace(r'\s+', '')
data1['產品品號']=data1['產品品號'].astype(str).replace(r'\s+', '')

# 檢查製令單的所有工單品號，在機種對照表中是否都存在
keep_rows = data1['產品品號'].isin(data2['途程品號'])
missing_rows = ~keep_rows

# 輸出缺失品號
if missing_rows.any():
  missing_products = data1[missing_rows]['產品品號'].tolist()
  print(f"缺失品號：{missing_products}")

# 从製令單中排除缺失品號
data1 = data1[keep_rows]



data1['類別'] = data1['產品品號'].apply(lambda x: data2[data2['途程品號'] == x]['類別'].values[0])
#sheet1的台/分去搜尋sheet2相同的產品品號那一列的暫時放寬11人(暫定工時)95%
data1['台/分'] = data1['產品品號'].apply(lambda x: data2[data2['途程品號'] == x]['暫時放寬11人\n(暫定工時)\n95%'].values[0])
time = data1['產量']/data1['台/分']
data1['加工時間'] = time.apply(math.ceil) #使用無條件進位,並套用到每個欄位中



# 将数据写入Excel文件
output_path = '製令單new.xlsx'
data1.to_excel(output_path, sheet_name='製令單', index=False)

# 打开Excel文件
workbook = openpyxl.load_workbook(output_path)
worksheet = workbook.active

# 调整 '製令編號' 列的列宽
for wid in ['A','B','C','D','E','F', 'G', 'H', 'I', 'J']:      #同時設定 F~K 欄寬為29
    worksheet.column_dimensions[wid].width = 20

#垂直水平置中,開啟自動換行
worksheet_alignment = Alignment(vertical='center', horizontal='center',wrapText=True)

#設定字型,文字大小,粗體,斜體
worksheet_font = Font(name='新細明體', size=12, bold=False, italic=False)

#將第1列到第36列,第1行到第11行,統一設定格式
for row in worksheet.iter_rows(min_row=1, max_row=40, min_col=1, max_col=21):
    for cell in row:
        cell.alignment = worksheet_alignment
        cell.font = worksheet_font

#=========================計算上班工時=========================================
row_names = sheet2.columns
wd = []
wd.append('日期')
for i in row_names:
    wd.append(i)


data= sheet1.iloc[:,0]
time = sheet1.iloc[:,4]
tm = []
day = []

#日期
for i in data:
    day.append(i)
del day[0]
#時數
for i in time:
    tm.append(i)
del tm[0]

#刪除空值
day = [x for x in day if x == x]

new_tm = []

for i in tm:
    if not pandas.isna(i):
        tm_values = [int(num) for num in str(i).split(',')]
        new_tm.append(tm_values)
    else:
        new_tm.append(i)  # 保留原本的NaN值

overtime = []  # 在外部定义overtime列表




a = 7
for i in range(len(wd)-1):
    overtime = sheet1.iloc[1:, a].tolist()
    for j in range(len(overtime)):
        if pandas.notna(overtime[j]):
            if isinstance(new_tm[j], list):
                new_tm[j][i] += overtime[j]
    a+=3

new_tm = [x for x in new_tm if x == x]


#表格的行列建立

df = pandas.DataFrame(new_tm)
df.insert(0,'日期',value=day)
df.columns = wd
df.to_excel('計算上班工時.xlsx')
#=====================================================

Changeover = pandas.read_excel("換線表測試_0117測試調整.xlsx")
classtime = pandas.read_excel("計算上班工時.xlsx")
Case = pandas.read_excel("製令單new.xlsx")
priority = pandas.read_excel("工時及可生產產品對應_V5.xlsx")

work_class = []
for i in sheet2.columns:
    work_class.append(i)

time = []
count_1 = 0  #排除前面的值 ex:新增工時、日期的值
count_2 = 0  #跳過加班時間
for i in sheet1.loc[1]:
    if count_1 > 4:
        count_2 += 1
        if count_2 %  3 != 0:
            time.append(i)
    count_1+=1

#剩餘時間
odd = 0

#補正時間
even = 1

a = 0
work_time=[]

#開始工作時間
while odd < len(time) and even < len(time):
    a = time[odd] + time[even]
    work_time.append(a)
    a = 0
    odd += 2
    even += 2

#------------------目前加工----------------------------
work_process = []
for i in sheet1.iloc[:0,~sheet1.columns.str.contains("^Unnamed")]:
    work_process.append(i)
del work_process[0]

dispatch = pandas.DataFrame(work_class)
dispatch.insert(1, column="開始工作", value=work_time)
dispatch.insert(2 ,column="目前加工類型", value=work_process)
dispatch.columns = ["index", "開始工作", "目前加工類型"]




#------------------計算出貨時間----------------------------

# 將日期轉換為想要的格式
classtime['日期'] = pandas.to_datetime(classtime['日期']).dt.strftime('%Y/%m/%d')
Case['預計出貨'] = pandas.to_datetime(Case['預計出貨']).dt.strftime('%Y/%m/%d')
classlen = len(sheet2.columns)
datalen = len(Case['預計出貨'])
sumtime = [[] for i in range(datalen)]

datacount = -1
Class = 0
classcount = 0
while datacount < datalen :
    datacount+=1
    if datacount == datalen:
        break
    ide = 0
    for i in classtime['日期']:
        ide = ide + 1
        if i == Case['預計出貨'][datacount]:
            break

    classcount = 1
    while classcount < classlen+1:
        sum = 0
        count = 0
        classcount+=1
        Class = classtime.columns[classcount]
        for i in classtime[Class]:
            if count == ide-1:
                break
            count += 1
            sum = sum + i
        sumtime[datacount].append(sum)




'''
#-----------------更新派工-----------------------------

input_class = input("輸入工班")
input_class = input_class + '班'
count = -1
for i in dispatch['index']:
    count += 1
    if i == input_class:
        break
value = 0
input_time = int(input("輸入時間"))
input_process = input('輸入加工類型')

value = dispatch.iat[count,1]
dispatch.at[count,"開始工作"] = input_time + value
dispatch.at[count,"目前加工類型"] = input_process
print(dispatch)


'''



#---------------開始派工-預計開工--------------------
start_construction = pandas.DataFrame()
sum_val = 0
counter = -1
number = -1

data_list = []

while sum_val <= len(work_class):
    number+=1

    if number == len(work_class):
        break
    for i in classtime[work_class[number]]:

        sum_val += i
        counter += 1
        if work_time[number] <= sum_val:
            date = classtime['日期'][counter]
            data_list.append(date)

            sum_val = 0  # 將 sum_val 的歸零移到這裡
            counter = -1  # 將 counter 的歸零移到這裡
            break
start_construction.insert(0, column="預計開工", value=data_list)

#---------------開始派工-預計完工--------------------

sum_val = 0
counter = -1
number = -1
date_end = []

while sum_val <= len(dispatch['開始工作']):
    number+=1
    if number == len(work_class):
        break
    for i in classtime[work_class[number]]:

        sum_val += i
        counter += 1
        if dispatch['開始工作'][number] <= sum_val:
            date = classtime['日期'][counter]
            date_end.append(date)
            sum_val = 0  # 將 sum_val 的歸零移到這裡
            counter = -1  # 將 counter 的歸零移到這裡
            break
start_construction.insert(1, column="預計完工", value=date_end)


#------------------換線時間-----------------
Changeover_dict = Changeover.to_dict()
row_dict = Changeover.iloc[1].to_dict()
column_dict = Changeover.iloc[:,1].to_dict()


counter =-1

processlen = len(Case['類別'])
line = [[] for i in range(processlen)]
lenprocess =len(work_process)
while counter <= len(Case['類別']):
    counter += 1

    if counter == len(Case['類別']):
        break
    classcount = -1

    while classcount <= lenprocess:
        x_count = -1
        y_count = -1
        classcount +=1
        if classcount == lenprocess:
            break
        for key, value in row_dict.items():
            x_count += 1
            if Case['類別'][counter] == value:

                break
        for key, value in column_dict.items():
            y_count += 1
            if dispatch['目前加工類型'][classcount] == value:
                break

        line[counter].append(Changeover.iloc[x_count,y_count])




#-----------------計算slt------------------------
addtime = [[] for i in range(len(line))]
sltime = [[] for i in range(len(line))]
counter = -1
each_wt = []
for i in Case['加工時間']:
    each_wt.append(i)

while counter < len(line):
    counter+=1
    classcount = -1
    if counter == len(line):
        break
    while classcount < classlen:
        classcount+=1
        if classcount ==classlen:
            break
        add = 0
        add = line[counter][classcount] + dispatch['開始工作'][classcount] + each_wt[counter]
        addtime[counter].append(add)
        sltime[counter].append(sumtime[counter][classcount]-add)

new_lst = pandas.DataFrame(sltime)
classnumber = ["製1班", "製2班", "製3班", "製4班", "製5班"]
new_lst.columns = classnumber[:classlen]

#-----------------各單優先權------------------------

priority_list = [[] for i in range(len(new_lst))]

priority_dict = priority.to_dict()
row = priority.loc[0].to_dict()
col = priority.iloc[1:,1].to_dict()



counter = -1
c = 0
r = -1
while counter <= len(new_lst):
    counter += 1
    if counter == len(new_lst):
        break
    classcount = -1
    while classcount <= classlen:
        classcount += 1
        r = -1
        c =  0
        if classcount == classlen:
            break
        for index, vaule in row.items():
            r = r + 1
            if classnumber[classcount] == vaule:
                break
        for index, vaule in col.items():
            c = c + 1
            if Case['類別'][counter] == vaule:
                break
        priority_list[counter].append(priority.iloc[c,r])

priority_df = pandas.DataFrame(priority_list)
priority_df.columns = classnumber[:classlen]



pandas.set_option('display.max_rows', None)
pandas.set_option('display.max_columns', None)

surface = new_lst
'''
storclass = []
for i in range(len(new_lst)):
    min = new_lst.idxmin().min()
    storclass.append(min)
    new_lst = new_lst.drop(min)
'''
priority_df['製1班'] = 1
priority_df['製2班'] = 1

#-----------------更新狀態表------------------------
one = [[] for i in range(len(new_lst))]
two = [[] for i in range(len(new_lst))]
three = [[] for i in range(len(new_lst))]
four = [[] for i in range(len(new_lst))]

one_count = 0
two_count = 0
three_count = 0
four_count = 0

def renewdipatch(select_class, select_index):

    global one_count
    global two_count
    global three_count
    global four_count

#===============換線時間===================
    x_count = -1
    y_count = -1

    mincase = Case['類別'][select_index]
    before_case = dispatch['目前加工類型'].iloc[select_class]
    for key, value in row_dict.items():
        x_count += 1
        if mincase == value:
            break
    for key, value in column_dict.items():
        y_count += 1
        if before_case == value:
            break

    new_line = Changeover.iloc[x_count,y_count]
    dispatch['開始工作'] = dispatch['開始工作'].astype(float)
    before_vaule=dispatch['開始工作'].iloc[select_class]+ new_line
    sum_vaule = before_vaule + each_wt[select_index]

    dispatch.at[select_class,"開始工作"] = sum_vaule
    dispatch.at[select_class,"目前加工類型"] = mincase

    sum_val = 0
    x_counter = -1

    for i in classtime[work_class[select_class]]:
        sum_val += i
        x_counter += 1
        if before_vaule <= sum_val:
            '''
            star_date = classtime['日期'][x_counter]
            '''
            break

    sum_val = 0
    y_counter = -1
    for i in classtime[work_class[select_class]]:
        sum_val += i
        y_counter += 1
        if sum_vaule <= sum_val:
            '''
            end_date = classtime['日期'][y_counter]
            '''
            break




    if(work_class[select_class] == '1班'):
        one[one_count].append(classtime['日期'][x_counter])
        one[one_count].append(classtime['日期'][y_counter])
        one[one_count].append(Case['預計出貨'][index_min])
        one[one_count].append(Case['產品品號'][index_min])
        one[one_count].append(Case['品名'][index_min])
        one[one_count].append(Case['規格'][index_min])
        one[one_count].append(Case['製令編號'][index_min])
        one[one_count].append(Case['產量'][index_min])
        one[one_count].append(Case['類別'][index_min])
        one[one_count].append(Case['台/分'][index_min])
        one[one_count].append(Case['加工時間'][index_min])
        one[one_count].append(new_line)


    elif(work_class[select_class] == '2班'):
        two[two_count].append(classtime['日期'][x_counter])
        two[two_count].append(classtime['日期'][y_counter])
        two[two_count].append(Case['預計出貨'][index_min])
        two[two_count].append(Case['產品品號'][index_min])
        two[two_count].append(Case['品名'][index_min])
        two[two_count].append(Case['規格'][index_min])
        two[two_count].append(Case['製令編號'][index_min])
        two[two_count].append(Case['產量'][index_min])
        two[two_count].append(Case['類別'][index_min])
        two[two_count].append(Case['台/分'][index_min])
        two[two_count].append(Case['加工時間'][index_min])
        two[two_count].append(new_line)


    elif(work_class[select_class] == '3班'):
        three[three_count].append(classtime['日期'][x_counter])
        three[three_count].append(classtime['日期'][y_counter])
        three[three_count].append(Case['預計出貨'][index_min])
        three[three_count].append(Case['產品品號'][index_min])
        three[three_count].append(Case['品名'][index_min])
        three[three_count].append(Case['規格'][index_min])
        three[three_count].append(Case['製令編號'][index_min])
        three[three_count].append(Case['產量'][index_min])
        three[three_count].append(Case['類別'][index_min])
        three[three_count].append(Case['台/分'][index_min])
        three[three_count].append(Case['加工時間'][index_min])
        three[three_count].append(new_line)


    elif(work_class[select_class] == '4班'):
        four[four_count].append(classtime['日期'][x_counter])
        four[four_count].append(classtime['日期'][y_counter])
        four[four_count].append(Case['預計出貨'][index_min])
        four[four_count].append(Case['產品品號'][index_min])
        four[four_count].append(Case['品名'][index_min])
        four[four_count].append(Case['規格'][index_min])
        four[four_count].append(Case['製令編號'][index_min])
        four[four_count].append(Case['產量'][index_min])
        four[four_count].append(Case['類別'][index_min])
        four[four_count].append(Case['台/分'][index_min])
        four[four_count].append(Case['加工時間'][index_min])
        four[four_count].append(new_line)


#-----------------更新換線------------------------
def renewline(list):
    for i in list:
       i.clear()
    counter = -1
    while counter <= len(Case['類別']):
        counter += 1
        if counter == len(Case['類別']):
            break
        classcount = -1
        while classcount <= lenprocess:
            x_count = -1
            y_count = -1
            classcount +=1
            if classcount == lenprocess:
                break
            for key, value in row_dict.items():
                x_count += 1
                if Case['類別'][counter] == value:
                    break
            for key, value in column_dict.items():
                y_count += 1
                if dispatch['目前加工類型'][classcount] == value:
                    break
            line[counter].append(Changeover.iloc[x_count,y_count])

#-----------------更新slacktime-----------------------
def renewslack(slt):
    for i in slt:
        i.clear()
    counter = -1
    while counter < len(line):
        counter+=1
        classcount = -1
        if counter == len(line):
            break
        while classcount < classlen:
            classcount+=1
            if classcount ==classlen:
                break
            add = int(line[counter][classcount]) + int(dispatch['開始工作'][classcount] + each_wt[counter])
            sltime[counter].append(sumtime[counter][classcount]-add)



#==============條件=============================


#proudct_index製令單編號
def rule(proudct_index):
    class_index = []
    count = 0
    max_value = priority_df.loc[proudct_index].max()

    for i in priority_df.loc[proudct_index].values:
        if i == max_value and sltime[proudct_index][count] >= 0:
            class_index.append(count)
        count+=1

    if len(class_index) < 2:
        return priority_df.loc[proudct_index].idxmax()

    Max = 0
    count = 0
    for i in class_index:
        if sltime[proudct_index][i] > Max:
            count = i
            Max = sltime[proudct_index][i]
    return priority_df.iloc[proudct_index].index[count]
#    return priority_df.iloc[proudct_index][count].columns


distribute = [[] for i in range(classlen)]
counter = -1
store = []
counter = 0
slt = []

while counter < len(surface):

    index_min = new_lst.idxmin().min()

    index_min_value = new_lst.min().min()

    slt = new_lst.loc[index_min].tolist()

    if counter == len(surface) :
        break

    Class = rule(index_min)
    store.append(index_min)
    # 移除最小的 index

    new_lst = new_lst.drop(index_min)

    if Class == '製1班':
#        one[one_count].append(index_min)
        renewdipatch(0, index_min)
        renewline(line)
        renewslack(sltime)

        one_count += 1

    elif Class == '製2班':
#        two[two_count].append(index_min)
        renewdipatch(1, index_min)
        renewline(line)
        renewslack(sltime)
        two_count += 1

    elif Class == '製3班':
#        three[three_count].append(index_min)
        renewdipatch(2, index_min)
        renewline(line)
        renewslack(sltime)

        three_count += 1

    elif Class == '製4班':
#        four[four_count].append(index_min)
        renewdipatch(3, index_min)
        renewline(line)
        renewslack(sltime)
        four_count += 1
    counter += 1

one_df = pandas.DataFrame(one)
two_df = pandas.DataFrame(two)
three_df = pandas.DataFrame(three)
four_df = pandas.DataFrame(four)

column = ['預計開工', '預計完工', '預計出貨', '產品品號', '品名', '規格', '製令編號', '產量', '類別', '台/分', '加工時間', '換線時間']


class_dfs =[]
if len(one[0]) != 0:
    one_df.columns= column
    class_dfs.append(one_df)
if len(two[0]) != 0:
    two_df.columns= column
    class_dfs.append(two_df)
if len(three[0]) != 0:
    three_df.columns= column
    class_dfs.append(three_df)
if len(four[0]) != 0:
    four_df.columns= column
    class_dfs.append(four_df)

####################################V換線優化V####################################

# 轉換日期欄位的格式並重新排列欄位順序
date_columns = ['預計開工', '預計完工', '預計出貨']  # 將 '預計完工' 加入日期欄位
i = 0
for i in range(len(class_dfs)):
    for col in date_columns:
        class_dfs[i][col] = pd.to_datetime(class_dfs[i][col])  # 只保留日期部分

# 要求使用者輸入每組的天數
group_days = int(input("請輸入每組的天數："))

# 將各工班的工單分組
grouped_orders = []
for class_df in class_dfs:
    # 將預計完工日期排序並重設索引
    class_df = class_df.sort_values('預計完工').reset_index(drop=True)

    # 初始化組別和日期計數器
    group = 1
    day_count = 0
    last_date = None

    # 遍歷每個工單
    for idx, row in class_df.iterrows():
        # 如果日期與上一個工單的日期不同，則日期計數器加一
        if row['預計完工'] != last_date:
            day_count += 1
            last_date = row['預計完工']

        # 如果日期計數器超過每組的天數，則組別加一並重設日期計數器
        if day_count > group_days:
            group += 1
            day_count = 1

        # 將工單的組別設為當前組別
        class_df.loc[idx, '組別'] = group

    # 如果迴圈結束後日期計數器小於每組的天數，則將剩下的工單都分到最後一組
    if day_count <= group_days:
        class_df.loc[class_df['組別'] > group, '組別'] = group

    # 將相同類別的工單放在一起
    class_df = class_df.sort_values(['組別', '類別', '預計出貨']).reset_index(drop=True)

    # 將 '組別' 列轉換為整數
    class_df['組別'] = class_df['組別'].astype(int)

    # 為所有訂單增加一項 "是否排序"，初始值為 False
    class_df['是否排序'] = False

    # 在 DataFrame 中添加一個新的列 "危險工單"，並將其預設值設為 False
    class_df['危險工單'] = False

    # 細分小組
    for group_num in class_df['組別'].unique():
        group_df = class_df[class_df['組別'] == group_num]
        latest_completion = group_df['預計完工'].max()
        subgroup1 = group_df[group_df['預計出貨'] <= latest_completion]
        subgroup2 = group_df[group_df['預計出貨'] > latest_completion]

        # 將 subgroup1 中的 "危險工單" 設為 True
        subgroup1.loc[:, '危險工單'] = True

        # 找出兩個小組中 "是否排序" 為 false 的類別的數量
        category_counts = pd.concat([subgroup1, subgroup2])[lambda x: x['是否排序'] == False]['類別'].value_counts()

        # 遍歷類別，直到找到一個在兩個小組中 "是否排序" 為 false 的類別，或者遍歷完所有的類別
        for category in category_counts.index:
            if category in subgroup1[subgroup1['是否排序'] == False]['類別'].values and category in subgroup2[subgroup2['是否排序'] == False]['類別'].values:
                common_category = category
                break
        else:
            common_category = None

        # 將當前小組中最多類別的訂單移到底部
        if len(subgroup1['類別'].unique()) > 1:
            sorted_subgroup1 = pd.concat([subgroup1[subgroup1['類別'] != common_category], subgroup1[subgroup1['類別'] == common_category]])
            sorted_subgroup1.loc[sorted_subgroup1['類別'] == common_category, '是否排序'] = True
        else:
            sorted_subgroup1 = subgroup1

        # 將下一個小組中最多類別的訂單移到頂部
        if len(subgroup2['類別'].unique()) > 1:
            sorted_subgroup2 = pd.concat([subgroup2[subgroup2['類別'] == common_category], subgroup2[subgroup2['類別'] != common_category]])
            sorted_subgroup2.loc[sorted_subgroup2['類別'] == common_category, '是否排序'] = True
        else:
            sorted_subgroup2 = subgroup2

        # #輸出第幾班哪個組別中最多的類別的訂單移到底部
        # print(f"第 {idx+1} 班：{group_num} 組中最多的類別的訂單移到底部"+'\n'+str(sorted_subgroup1) + '\n')
        # #輸出第幾班哪個組別中最多的類別的訂單移到頂部
        # print(f"第 {idx+1} 班：{group_num} 組中最多的類別的訂單移到頂部"+'\n'+str(sorted_subgroup2) + '\n')

        # 移除 class_df 中當前組別的資料
        class_df = class_df[class_df['組別'] != group_num]

        # 將 sorted_subgroup1 和 sorted_subgroup2 連接起來
        sorted_group = pd.concat([sorted_subgroup1, sorted_subgroup2])

        # 創建一個新的索引，從當前 group_df 的最小索引開始
        new_index = range(group_df.index.min(), group_df.index.min() + len(sorted_group))

        # 將新的索引賦值給 sorted_group
        sorted_group.index = new_index

        # 將排序後的資料加入到 class_df 中
        class_df = pd.concat([class_df, sorted_group])

        # # 輸出更新到 class_df 的小組
        # print("更新到 class_df 的小組：")
        # print(class_df.loc[class_df['組別'] == group_num])

    # 將分組的工單加入到結果中
    grouped_orders.append(class_df)

# 輸出分組的工單
for i, grouped_order in enumerate(grouped_orders, start=1):
    print(f"第 {i} 班：")
    for group, group_df in grouped_order.groupby('組別'):
        print(f"第 {group} 組：")
        print(group_df)
        print("-"*100)
    print("="*200)
    print("="*200)

# 處理各班的各小組間
for idx, class_df in enumerate(grouped_orders):

    # 獲取所有的組別
    groups = class_df['組別'].unique()
    
    #輸出第幾班有幾個組別
    print(f"第 {idx+1} 班有 {len(groups)} 個組別")
    
    #將預計
    
    # 遍歷每個組別
    for group_index in range(len(groups) - 1):
        
        #輸出組別
        print(f"第 {idx+1} 班：{groups[group_index]} 組和 {groups[group_index+1]} 組"+'\n')
        
        # 獲取當前組別和下一個組別
        group1 = class_df.loc[(class_df['組別'] == groups[group_index]) & (class_df['危險工單'] == False)]
        if group1.empty or group1['危險工單'].all():
            group1 = class_df.loc[class_df['組別'] == groups[group_index]]

        group2 = class_df.loc[(class_df['組別'] == groups[group_index+1]) & (class_df['危險工單'] == True)]
        if group2.empty or group2['危險工單'].all() == False:
            group2 = class_df.loc[class_df['組別'] == groups[group_index+1]]

        # 找出兩個組別中 "是否排序" 為 false 的類別的數量
        category_counts = pd.concat([group1, group2])[lambda x: x['是否排序'] == False]['類別'].value_counts()

        # 遍歷類別，直到找到一個在兩個組別中 "是否排序" 為 false 的類別，或者遍歷完所有的類別
        for category in category_counts.index:
            if category in group1[group1['是否排序'] == False]['類別'].values and category in group2[group2['是否排序'] == False]['類別'].values:
                common_category = category
                break
        else:
            common_category = None

        if common_category is None:
            continue

        #輸出第幾班哪兩個組別中最多的類別
        print(f"第 {idx+1} 班：{groups[group_index]} 組和 {groups[group_index+1]} 組中最多的類別為 {common_category}")

        # 將當前組別中最多類別的訂單移到底部
        if len(group1['類別'].unique()) > 1:
            sorted_group1 = pd.concat([group1[group1['類別'] != common_category], group1[group1['類別'] == common_category]])
            sorted_group1.loc[sorted_group1['類別'] == common_category, '是否排序'] = True
        else:
            sorted_group1 = group1
            sorted_group1.loc[sorted_group1['類別'] == common_category, '是否排序'] = True
        #輸出第幾班哪個組別中最多的類別的訂單移到底部
        print(f"第 {idx+1} 班：{groups[group_index]} 組中最多的類別的訂單移到底部"+'\n'+str(sorted_group1) + '\n')

        # 將下一個組別中最多類別的訂單移到頂部
        if len(group2['類別'].unique()) > 1:
            sorted_group2 = pd.concat([group2[group2['類別'] == common_category], group2[group2['類別'] != common_category]])
            sorted_group2.loc[sorted_group2['類別'] == common_category, '是否排序'] = True
        else:
            sorted_group2 = group2
        #輸出第幾班哪個組別中最多的類別的訂單移到底部
        print(f"第 {idx+1} 班：{groups[group_index+1]} 組中最多的類別的訂單移到頂部"+'\n'+str(sorted_group2) + '\n')
        
        # 移除 class_df 中當前組別的資料
        if group1.empty or group1['危險工單'].all():
            class_df = class_df.loc[class_df['組別'] != groups[group_index]]
        else:
            class_df = class_df[~((class_df['組別'] == groups[group_index]) & (class_df['危險工單'] == False))]
            
        if group2.empty or group2['危險工單'].all() == False:
            class_df = class_df.loc[class_df['組別'] != groups[group_index+1]]
        else:
            class_df = class_df[~((class_df['組別'] == groups[group_index+1]) & (class_df['危險工單'] == True))]
        
        # 將 sorted_group1 和 sorted_group2 連接起來
        sorted_group = pd.concat([sorted_group1, sorted_group2])
        
        # 創建一個新的索引，從當前 group_df 的最小索引開始
        new_index = range(group1.index.min(), group1.index.min() + len(sorted_group))
        
        # 將新的索引賦值給 sorted_group
        sorted_group.index = new_index
        
        # 將排序後的資料加入到 class_df 中
        class_df = pd.concat([class_df, sorted_group])

    # #輸出第幾班最後的結果
    # print(f"第 {idx+1} 班最後的結果"+'\n'+str(class_df) + '\n')

    # 更新 grouped_orders
    grouped_orders[idx] = class_df

    #print("="*100)

# 輸出分組的工單
for i, grouped_order in enumerate(grouped_orders, start=1):
    print(f"第 {i} 班：")
    for group, group_df in grouped_order.groupby('組別'):
        print(f"第 {group} 組：")
        print(group_df)
        print("-"*100)
    print("="*200)
    print("="*200)

#刪除組別和是否排序和危險工單
for idx, class_df in enumerate(grouped_orders):
    class_df = class_df.drop(columns=['組別', '是否排序', '危險工單'])
    grouped_orders[idx] = class_df

#class_dfs更新成grouped_orders
class_dfs = grouped_orders

def update_changeover_time(class_dfs, Changeover, sheet1):
    for i, class_df in enumerate(class_dfs):
        start_work_time=0
        end_work_time=0
        # 初始化開始工作時間為0
        time_count = 0
        j = 0
        for j in range(len(class_df)):
            if j == 0:  # 第一張工單
                prev_order = sheet1.columns[i*3+5]
                # 從 sheet1 的第二行中獲取開始工作時間
                time_count = sheet1.iloc[1, i*3+5] + sheet1.iloc[1, i*3+6]
            else:  # 其他工單
                prev_order = class_df.iloc[j-1]['類別']

            current_order = class_df.iloc[j]['類別']

            # #輸出前一張工單和當前工單
            # print(f"第 {i+1} 班：第 {j+1} 張工單：前一張工單為 {prev_order}，當前工單為 {current_order}")

            # #輸出Changeover第一列
            # print(f"第 {i+1} 班：Changeover第一列"+'\n'+str(Changeover.iloc[1]) + '\n')

            # #輸出Changeover第一行
            # print(f"第 {i+1} 班：Changeover第一行"+'\n'+str(Changeover.iloc[:, 1]) + '\n')

            # 在第一列中獲取 prev_order 的列索引
            prev_order_index = Changeover.iloc[1].tolist().index(prev_order)

            # 在第一行中獲取 current_order 的行索引
            current_order_index = Changeover.iloc[:, 1].tolist().index(current_order)

            # 從 Changeover 表找到兩類別十字交叉的儲存格數字為換線時間
            changeover_time = Changeover.iloc[current_order_index, prev_order_index]

            # 更新換線時間
            class_df.at[j, '換線時間'] = changeover_time

            #time_count加上changeover_time
            time_count += changeover_time
            start_work_time = time_count

            #輸出工班工單開工時間
            print(f"第 {i+1} 班 第 {j+1} 單：工班工單開工時間"+'\n'+str(start_work_time) + '\n')

            # 從 classtime 表找到對應的開始日期
            start_date = classtime.iloc[(classtime.iloc[:, i+2].cumsum() > start_work_time).idxmax(), 1]
            # 更新預計開工日期
            class_df.at[j, '預計開工'] = start_date


            time_count += class_df.iloc[j]['加工時間']
            end_work_time = time_count

            #輸出工班工單完工時間
            print(f"第 {i+1} 班 第 {j+1} 單：工班工單完工時間"+'\n'+str(end_work_time) + '\n')

            # 從 classtime 表找到對應的結束日期
            end_date = classtime.iloc[(classtime.iloc[:, i+2].cumsum() >= end_work_time).idxmax(), 1]
            # 更新預計完工日期
            class_df.at[j, '預計完工'] = end_date

    return class_dfs

# 更新換線時間
class_dfs = update_changeover_time(class_dfs, Changeover, sheet1)

####################################V輸出排程表V####################################

def create_excel(class_dfs,classtime,sheet1,waiting_time_sheet):
    # 建立新的 Excel 檔案
    output_workbook = Workbook()

    # 刪除第一個空白的工作表
    output_workbook.remove(output_workbook.active)

    # 讀取'待填工時表.xlsx'並轉換成 DataFrame
    waiting_time_sheet['日期'] = pd.to_datetime(waiting_time_sheet['日期'])

    # 根據日期範圍分組
    grouped_dates = waiting_time_sheet.groupby('日期')['星期'].first().reset_index()
    grouped_dates['週起始日期'] = grouped_dates['日期'] - pd.to_timedelta(grouped_dates['日期'].dt.dayofweek, unit='d')
    grouped_dates['週結束日期'] = grouped_dates['週起始日期'] + pd.to_timedelta(5, unit='d')
    grouped_dates = grouped_dates.groupby(['週起始日期', '週結束日期']).agg({'日期': list}).reset_index()

    # 創建一個字典來記錄每個班級的 "(本)新增工時"
    remaining_work_time_dict = {}

    # 寫入各週的製令單到不同的工作表中
    for idx, row in grouped_dates.iterrows():
        start_date = row['週起始日期'].strftime('%Y-%m-%d')
        end_date = row['週結束日期'].strftime('%Y-%m-%d')

        # 檢查是否有任一班次有被分配到工單
        has_assigned_work_for_week = any(class_df['預計開工'].between(start_date, end_date).any() for class_df in class_dfs)

        if has_assigned_work_for_week:
            week_sheet = output_workbook.create_sheet(title=f"{start_date} ~ {end_date}")

            current_row = 1

            # 檢查各班次是否都有工單
            for class_idx, class_df in enumerate(class_dfs):
                has_assigned_work_for_class = class_df['預計開工'].between(start_date, end_date).any()

                if has_assigned_work_for_class:
                    # 寫入班次標題
                    week_sheet.merge_cells(start_row=current_row, end_row=current_row + 1, start_column=1, end_column=3)
                    week_sheet.cell(row=current_row, column=1, value=f"製造 {class_idx+1} 班").font = Font(size=24)
                    current_row += 2

                    # 分成三組資料
                    grouped_common_data = [
                        {'稼動總工時': None, '損耗率(%)': None, '製令別': '510廠內製令'},
                        {'(上)剩餘工時': None, '換線補正': None, '補正工時': None, '製表日期': None},
                        {'(本)新增工時': None, '(本)剩餘工時': None, '加班工時': None, '備註': None}
                    ]

                    # 將 classtime 中的日期轉換成日期格式
                    classtime.iloc[:, 1] = pd.to_datetime(classtime.iloc[:, 1])

                    # 將開始日期和結束日期轉換成日期格式
                    start_date = pd.to_datetime(start_date)
                    end_date = pd.to_datetime(end_date)

                    # 計算稼動總工時，按照當週的日期範圍，在 classtime 中找到對應的開始時間和結束時間，然後相加
                    total_work_time = 0
                    i = 0
                    for i in range(len(classtime)):
                        if classtime.iloc[i, 1] >= start_date:
                            if classtime.iloc[i, 1] > end_date:
                                break
                            total_work_time += classtime.iloc[i, class_idx+2]
                    grouped_common_data[0]['稼動總工時'] = total_work_time

                    #損耗率讀取Loss rate.txt
                    with open(Loss_rate, 'r') as f:
                        loss_rate = f.read()
                        #根據班級索引值選取讀取第幾行
                        loss_rate = loss_rate.split('\n')[class_idx]
                        # 將讀取的損耗率轉換成浮點數
                        loss_rate = float(loss_rate)
                        loss_rate = str(loss_rate)
                        #輸出損耗率
                        print(f"第 {class_idx+1} 班：損耗率為 {loss_rate}")
                    grouped_common_data[0]['損耗率(%)'] = loss_rate + "%"

                    #計算(上)剩餘工時
                    last_work_time = 0
                    if idx == 0:
                        last_work_time = sheet1.iloc[1, class_idx*3+5]
                    else:
                        last_work_time = remaining_work_time_dict[f"班級{class_idx+1}"] * -1
                    grouped_common_data[1]['(上)剩餘工時'] = last_work_time

                    # 計算換線補正，將 class_df "預計開工" 日期在當週範圍內的換線時間相加
                    changeover_correction = 0
                    changeover_correction = class_df.loc[class_df['預計開工'].between(start_date, end_date), '換線時間'].sum()
                    grouped_common_data[1]['換線補正'] = changeover_correction

                    # 計算補正工時第一週讀取sheet1的第一列，其他為0
                    correction_time = 0
                    if idx == 0:
                        correction_time = sheet1.iloc[1, class_idx*3+6]
                    grouped_common_data[1]['補正工時'] = correction_time

                    #製表日期使用當前日期
                    table_date = datetime.now().strftime('%Y/%m/%d')
                    grouped_common_data[1]['製表日期'] = table_date

                    #(本)新增工時
                    new_work_time = 0
                    new_work_time = class_df.loc[class_df['預計開工'].between(start_date, end_date), '加工時間'].sum()
                    grouped_common_data[2]['(本)新增工時'] = new_work_time

                    #剩餘工時
                    remaining_work_time = 0
                    remaining_work_time = total_work_time - last_work_time - changeover_correction - correction_time - new_work_time
                    grouped_common_data[2]['(本)剩餘工時'] = remaining_work_time
                    #將(本)剩餘工時記錄在new_work_time_dict
                    remaining_work_time_dict[f"班級{class_idx+1}"] = remaining_work_time

                    #加班工時
                    overtime = 0
                    ew_work_time = waiting_time_sheet.loc[waiting_time_sheet['日期'].between(start_date, end_date), f'製造{class_idx+1}班\n加班工時(小時)'].sum()
                    grouped_common_data[2]['加班工時'] = ew_work_time

                    # 寫入共同資料到 DataFrame
                    for i, common_group in enumerate(grouped_common_data):
                        for j, (key, value) in enumerate(common_group.items(), start=1):
                            key = f"{key}:"
                            if i == 0 and j == 3:
                                cell = week_sheet.cell(row=current_row + i, column=j * 2 + 1, value=key)
                                cell = week_sheet.cell(row=current_row + i, column=j * 2 + 2, value=value)
                            else:
                                cell = week_sheet.cell(row=current_row + i, column=j * 2 - 1, value=key)
                                cell = week_sheet.cell(row=current_row + i, column=j * 2, value=value)
                            if key == '(本)剩餘工時:' and value is not None and value > 0:
                                cell.fill = PatternFill(start_color="FFE66F", end_color="FFE66F", fill_type="solid")

                    # 移動到下一行
                    current_row += len(grouped_common_data) + 1

                    # 寫入 DataFrame 標題
                    for col_idx, col_name in enumerate(class_df.columns, start=1):
                        cell = week_sheet.cell(row=current_row, column=col_idx, value=col_name)
                        cell.font = Font(bold=True)
                        cell.border = Border(left=Side(style='thin', color='000000'),
                                            right=Side(style='thin', color='000000'),
                                            top=Side(style='thin', color='000000'),
                                            bottom=Side(style='thin', color='000000'))
                        # 設定欄寬
                        if col_idx in [1, 2, 3, 4, 5, 7, 8, 9, 10]:
                            week_sheet.column_dimensions[cell.column_letter].width = 16.29
                        elif col_idx == 6:
                            week_sheet.column_dimensions[cell.column_letter].width = 27.29
                        elif col_idx in [11, 12]:
                            week_sheet.column_dimensions[cell.column_letter].width = 10

                    current_row += 1

                    # 寫入 DataFrame 資料
                    for _, row in class_df.iterrows():
                        # 檢查該工單是否在這週的日期範圍內
                        if row['預計開工'] >= start_date and row['預計開工'] <= end_date:
                            # 檢查 '預計完工日期' 是否大於等於 '預計出貨日期'
                            is_late = row['預計完工'] >= row['預計出貨']
                            for col_idx, value in enumerate(row, start=1):
                                if isinstance(value, pd.Timestamp):
                                    value = value.strftime('%Y/%m/%d')
                                cell = week_sheet.cell(row=current_row, column=col_idx, value=value)
                                # 如果 '預計完工日期' 大於等於 '預計出貨日期'，則將單元格設為黃色
                                if is_late:
                                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            current_row += 1

                    # 插入空行分隔各班次
                    current_row += 1

                    # 設定工作表比例
                    week_sheet.sheet_view.zoomScale = 135

    # 將所有儲存格文字靠左
    for sheet in output_workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

    # 儲存新的 Excel 檔案
    output_workbook.save('排程表.xlsx')

create_excel(class_dfs,classtime,sheet1,waiting_time_sheet)
